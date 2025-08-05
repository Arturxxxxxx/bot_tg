from collections import defaultdict
from datetime import date, datetime, timedelta
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from slugify import slugify

from data.database import conn

DAYS_RU = ["Понедельник", "Вторник", "Среда",
           "Четверг", "Пятница", "Суббота", "Воскресенье"]


# ---------- USER REPORT --------------------------------------------------
def generate_user_excel(user_id: int, company_name: str, monday: date) -> str:
    """
    Формирует Excel‑отчёт компании за неделю, начинающуюся с monday (ISO‑дата понедельника).
    Возвращает локальный путь к файлу.
    """
    safe = slugify(company_name or user_id)
    sunday = monday + timedelta(days=6)
    fname  = f"exports/user_{safe}_{monday:%d-%m}_{sunday:%d-%m}.xlsx"
    os.makedirs("exports", exist_ok=True)

    cur = conn.cursor()
    cur.execute(
        """
        SELECT day, time, portion, created_at
        FROM   portions
        WHERE  user_id=? AND day BETWEEN ? AND ?
        ORDER  BY day, time
        """,
        (user_id, monday.isoformat(), sunday.isoformat()),
    )
    rows = cur.fetchall()                                    # (iso_day, time, portion, created)

    # (day,time) -> (portion, created_at)
    m = {(d, t): (p, c) for d, t, p, c in rows}

    # wb, ws = Workbook(), Workbook().active
    wb = Workbook()
    ws = wb.active 
    ws.title = "Заявки компании"

    # ---- стили -----------------------------------------------------------
    bold  = Font(bold=True)
    hfill = PatternFill("solid", start_color="D9E1F2")
    tfill = PatternFill("solid", start_color="BDD7EE")
    border = Border(*(Side("thin"),)*4)

    # ---- шапка -----------------------------------------------------------
    ws.merge_cells("A1:D1")
    ws["A1"].value, ws["A1"].font, ws["A1"].alignment, ws["A1"].fill = (
        "Заявки на питание", Font(size=14, bold=True), Alignment("center"), tfill
    )
    ws["A1"].border = border

    ws.merge_cells("A2:D2")
    ws["A2"].value, ws["A2"].font = f"Компания: {company_name}", Font(size=12, bold=True)
    ws["A2"].border = border

    headers = ["День", "Время", "Порции", "Дата заявки"]
    ws.append(headers)
    for c in ws[3]:
        c.font, c.alignment, c.fill, c.border = bold, Alignment("center"), hfill, border

    # ---- данные ----------------------------------------------------------
    times = ["День", "Ночь", "Выпечка"]
    for i in range(7):
        d = monday + timedelta(days=i)
        iso = d.isoformat()
        for t in times:
            portion, created = m.get((iso, t), (0, "—"))
            ws.append([iso, t, portion, created])
            for cell in ws[ws.max_row]:
                cell.alignment, cell.border = Alignment("center"), border

    # ---- автоширина ------------------------------------------------------
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            len(str(c.value)) if c.value else 0 for c in col
        ) + 2

    wb.save(fname)
    return fname


# ---------- ADMIN REPORT -------------------------------------------------
def generate_admin_excel(year: int, week_num: int) -> str:
    """
    Формирует общий отчёт за ISO‑неделю year‑Wweek_num (понедельник‑воскресенье)
    по всем компаниям. Файл сохраняется в ./exports и возвращается его путь.
    """
    monday  = datetime.fromisocalendar(year, week_num, 1).date()
    sunday  = monday + timedelta(days=6)
    fname   = f"exports/admin_orders_{year}-W{week_num:02d}.xlsx"
    os.makedirs("exports", exist_ok=True)

    # ---- Читаем БД -------------------------------------------------------
    cur = conn.cursor()
    cur.execute(
        """
        SELECT company_name,
               day,                -- ISO‑дата (TEXT)
               time,               -- 'День' | 'Ночь' | 'Запайка'
               SUM(portion)        -- суммуем сразу
        FROM   portions
        WHERE  day BETWEEN ? AND ?
        GROUP  BY company_name, day, time
        """,
        (monday.isoformat(), sunday.isoformat()),
    )
    rows = cur.fetchall()            # (comp, iso_day, time, sum_portion)

    # company → {(iso_day, time) → portion}
    data: dict[str, dict[tuple[str, str], int]] = defaultdict(dict)
    for comp, iso_d, t, total in rows:
        data[comp][(iso_d, t.capitalize())] = total

    # ---- Создаём Excel ---------------------------------------------------
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"W{week_num:02d}"

    # Стили
    bold   = Font(bold=True)
    tfill  = PatternFill("solid", start_color="BDD7EE")
    hfill  = PatternFill("solid", start_color="D9E1F2")
    border = Border(*(Side("thin"),)*4)

    # Заголовок файла
    ws.merge_cells("A1:F1")
    top = ws["A1"]
    top.value = f"Календарь заявок {year}-W{week_num:02d} " \
                f"({monday:%d.%m}–{sunday:%d.%m})"
    top.font, top.alignment, top.fill = Font(size=14, bold=True), Alignment("center"), tfill
    top.border = border

    # Шапка таблицы
    ws.append(["Компания", "Дата", "День", "Ночь", "Выпечка", "Итого"])
    for c in ws[2]:
        c.font, c.alignment, c.fill, c.border = bold, Alignment("center"), hfill, border

    times = ["День", "Ночь", "Выпечка"]
    row_i = 3

    for comp, recs in data.items():
        for i in range(7):
            d        = monday + timedelta(days=i)
            iso_d    = d.isoformat()
            portions = [recs.get((iso_d, t), 0) for t in times]
            total    = sum(portions)

            ws.append([comp, iso_d, *portions, total])
            for c in ws[row_i]:
                c.alignment, c.border = Alignment("center"), border
            row_i += 1

        # пустая разделительная строка
        ws.append([])
        row_i += 1

    # Авто‑ширина
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            max(len(str(c.value)) if c.value else 0 for c in col) + 2
        )

    wb.save(fname)
    return fname